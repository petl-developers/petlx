"""
Read and write xlsx files, using openpyxl.

"""

import os

import petl

from petlx.util import UnsatisfiedDependency


dep_message = """
The package openpyxl is required. Instructions for installation can be found at 
https://bitbucket.org/ericgazoni/openpyxl/wiki/Home or try pip install openpyxl.
"""


def fromxlsx(filename, sheet=None, range_string=None, row_offset=0,
             column_offset=0, **kwargs):
    """
    Extract a table from a sheet in an Excel (.xlsx) file.
    
    N.B., the sheet name is case sensitive, so watch out for, e.g., 'Sheet1'.

    .. versionchanged:: 0.15

    The ``sheet`` argument can be omitted, in which case the first sheet in
    the workbook is used by default.

    The ``range_string`` argument can be used to provide a range string
    specifying a range of cells to extract.

    The ``row_offset`` and ``column_offset`` arguments can be used to
    specify offsets.

    Any other keyword arguments are passed through to
    :func:`openpyxl.load_workbook()`.

    .. versionchanged:: 0.18

    The keyword argument 'range' has been renamed to 'range_string' for
    consistency with the underlying openpyxl API.

    """

    # backwards compatibility
    if 'range' in kwargs:
        range_string = kwargs.pop('range')

    return XLSXView(filename, sheet=sheet, range_string=range_string,
                    row_offset=row_offset, column_offset=column_offset,
                    **kwargs)


class XLSXView(petl.util.RowContainer):
    
    def __init__(self, filename, sheet=None, range_string=None,
                 row_offset=0, column_offset=0, **kwargs):
        self.filename = filename
        self.sheet = sheet
        self.range_string = range_string
        self.row_offset = row_offset
        self.column_offset = column_offset
        self.kwargs = kwargs

    def __iter__(self):
        try:
            import openpyxl
        except ImportError as e:
            raise UnsatisfiedDependency(e, dep_message)

        wb = openpyxl.load_workbook(filename=self.filename,
                                    use_iterators=True, **self.kwargs)
        if self.sheet is None:
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        elif isinstance(self.sheet, int):
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[self.sheet])
        else:
            ws = wb.get_sheet_by_name(str(self.sheet))

        return (tuple(cell.value for cell in row)
                for row in ws.iter_rows(range_string=self.range_string,
                                        row_offset=self.row_offset,
                                        column_offset=self.column_offset))


def toxlsx(tbl, filename, sheet=None, encoding='utf-8'):
    """
    Write a table to a new Excel (.xlsx) file.

    .. versionadded:: 0.15

    """

    try:
        import openpyxl
    except ImportError as e:
        raise UnsatisfiedDependency(e, dep_message)
    wb = openpyxl.Workbook(optimized_write=True, encoding=encoding)
    ws = wb.create_sheet(title=sheet)
    for row in tbl:
        ws.append(row)
    wb.save(filename)


import sys
from petlx.integration import integrate
integrate(sys.modules[__name__])

